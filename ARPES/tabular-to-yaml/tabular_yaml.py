import pandas as pd
import numpy as np
import yaml
import pprint
import argparse
import re
import os
# from pandas.core.dtypes.missing import isnull
from pandas import ExcelWriter
from openpyxl import load_workbook

"""
Tabular-YAML two-way conversion script for ARPES definitions
Required arguments: -n <filename> -d <destination folder>

Example conversion from xlsx to yaml:
    ./tabular_yaml.py -n mpes-nexus_metadata_parameters_reviewed.xlsx -d testing

Example conversion from yaml to xlsx:
    ./tabular_yaml.py -n testing/NXsample_converted.yaml -d to_xlsx
"""


def csv_to_yaml():
    print("Not fully tested")
    yaml_filename = args.filename.replace(args.filename.split(".", 1)[1], 'yml')
    temp_df = pd.read_csv(args.filename)
    temp_df.dropna(inplace=True)
    df = temp_df.drop_duplicates(subset=['Name'], inplace=False)
    diff = len(temp_df)-len(df)
    if(diff != 0):
        print("WARNING: One or more column(s) had the same name and only the first was kept.")
        print("     Number of rows dropped:" + str(diff))

    df.set_index('Name', inplace=True)
    sheet_dict = df.to_dict('index')
    with open(yaml_filename, "w") as f:
        yaml.dump(sheet_dict, f)


def stripspaces(string):
    return re.sub(r"\s+", "", str(string), flags=re.UNICODE)


def cleanDF(df):
    df = df.fillna('')
    df['Nexus hierarchy'] = df['Nexus hierarchy'].str.split(':').str[-1]
    MergedTypeClass = df['type:units'].str.split(':').str[0].apply(stripspaces) + df['NX class'].astype(str).apply(stripspaces)
    df['Name'] = df['Nexus hierarchy'].apply(stripspaces) + "(" + MergedTypeClass + ")"
    df = df[df.Name != "()"]
    units = [stripspaces(s.split(':')[-1]) if len(s.split(':')) > 1 else "" for s in df['type:units']]
    df['unit'] = units
    df['doc'] = "\"" + df['Documentation'].replace('', np.nan, inplace=False) + "\""
    df.drop(columns=["Nexus hierarchy", "type:units", "NX class", "Documentation"], inplace=True)
    df.set_index('Name', inplace=True)
    df.dropna(axis=0, thresh=3, inplace=True)
    df.dropna(axis=1, thresh=3, inplace=True)
    check_dupes = df.index.duplicated(keep='first')
    num = 0
    while any(check_dupes):
        new_indices = [str(s) if not check_dupes[ind] else (str(s) + "_dupe" + str(num)) for ind, s in enumerate(df.index)]
        num += 1
        df.index = new_indices
        check_dupes = df.index.duplicated(keep='first')

    return df


def clean_duplicate_handler(filepath):
    with open(filepath, "r", errors='ignore') as text_file:
        data = text_file.read()
        # print(re.findall(r'_dupe\w+',data))
        data = re.sub(r'_dupe\w+', "", data)
        data = re.sub(r'\'\[\[', "[[", data)
        data = re.sub(r'\]\]\'', "]]", data)
        data = re.sub(r'\'\"', "\"", data)
        data = re.sub(r'\"\'', "\"", data)

    with open(filepath, "w") as text_file:
        text_file.write(data)

    text_file.close()


def appendDict(keys, value, dict):
    if len(keys) == 1:
        dict[keys[0].strip()] = value
    else:
        if keys[0] not in dict:
            dict[keys[0]] = {}
        appendDict(keys[1:], value, dict[keys[0]])


def xlsx_to_yaml():
    xls = load_workbook(file_path, read_only=True, data_only=True)
    names = xls.sheetnames
    # names = ["NXsample"]

    for sheet in names:
        print("Processing sheet: " + sheet)
        yaml_filename = sheet + ".yml"
        ws = xls[sheet]
        temp_df = pd.DataFrame(ws.values, columns=next(ws.values)[0:])
        df = temp_df.drop_duplicates(subset=['Nexus hierarchy'], inplace=False).iloc[1:, :]
        df = df.loc[df['Exists'] == "F"]
        if(df.empty):
            continue
        diff = len(temp_df)-len(df)

        if(diff != 0):
            print("WARNING: One or more column(s) had the same nexus hierarchy and only the first was kept.\nThis can happen if your file has merged rows.")
            print("     Sheet:" + sheet)
            print("     Number of rows dropped:" + str(diff))
        
        df = df.fillna('')
        MergedTypeClass = df['type:units'].str.split(':').str[0].apply(stripspaces) + df['NX class'].astype(str).apply(stripspaces)
        df = df[df.Name != ""]
        df['Nexus hierarchy'] = df['Nexus hierarchy'].str.split(':')
        df['Nexus hierarchy'] = [[str(s).strip() + "(" + MergedTypeClass.iloc[index] + ")" for s in temp] for index, temp in enumerate(df['Nexus hierarchy'])]
        df['Count'] = [len(temp) if temp else np.nan for temp in df['Nexus hierarchy']]
        units = [stripspaces(s.split(':')[-1]) if len(s.split(':')) > 1 else "" for s in df['type:units']]
        df['unit'] = units
        df['doc'] = "\"" + df['Documentation'].replace('', np.nan, inplace=False) + "\""
        print(df['Count'].mode())
        mode = int(df['Count'].mode())
        output = {}
        for index, row in df.iterrows():
            data = {'doc': str(row['doc']),
                    'unit': row['unit'],
                    'Dim': row['Dim'],
                    'Rank': row['Rank']}
            appendDict(row['Nexus hierarchy'][mode-1:], data, output)

        sheet_dict = {
            outer_k: {
                inner_k: inner_v
                for inner_k, inner_v in outer_v.items()
                if inner_v != 0 and inner_v != ""
            }
            for outer_k, outer_v in output.items()
        }

        for key in sheet_dict:
            if "Dim" in sheet_dict.get(key):
                dim = sheet_dict.get(key).pop("Dim")
                rank = sheet_dict.get(key).pop("Rank")
                if dim != 0:
                    dim = "[" + str(dim) + "]"
                    newitem = {"dimensions": {"dim": dim, "rank": rank}}
                    sheet_dict.get(key).update(newitem)

        pprint.pprint(sheet_dict)
        output_path = os.path.join(folder_path, yaml_filename)
        output = {}
        with open("Headers.yml", 'r') as stream:
            try:
                parsed_yaml = yaml.safe_load(stream)
                if(sheet in parsed_yaml):
                    if 'doc' in parsed_yaml[sheet]:
                        parsed_yaml[sheet].update({'doc': "\"" + str(parsed_yaml[sheet]['doc']) + "\""})
                    output.update(parsed_yaml[sheet])
                    output['(NXobject)'] = sheet_dict
                else:
                    print("ERROR: Not found in Headers.yml; defaulting to no header information.")
                    tempstring = "(" + sheet + ")"
                    output[tempstring] = sheet_dict
            except yaml.YAMLError as exc:
                print(exc)
        with open(output_path, "w") as f:
            yaml.dump(output, f, sort_keys=False)
            f.close()

        clean_duplicate_handler(output_path)

    xls.close()
    print("The new file has been saved as " + yaml_filename)

            # count = int(mode)-1
            # print(row['Count'])
            # while count<int(row['Count']):
            #     head = row['Nexus hierarchy'][count]
            #     print(head)
            #     # if count == int(row['Count'])-1:
            #     #     head = {}
            #     # else:
            #     #     output[row['Nexus hierarchy'][count]]
            #     count += 1


# def xlsx_to_yaml():
#     # xls = pd.ExcelFile(file_path,engine='openpyxl')
#     xls = load_workbook(file_path, read_only=True, data_only=True)
#     names = xls.sheetnames
#     # names = ["NXsample"]

#     for sheet in names:
#         print("Processing sheet: " + sheet)
#         # yaml_filename = sheet + "_converted.yaml"
#         yaml_filename = sheet + ".yml"
#         ws = xls[sheet]
#         temp_df = pd.DataFrame(ws.values, columns=next(ws.values)[0:])
#         df = temp_df.drop_duplicates(subset=['Nexus hierarchy'], inplace=False).iloc[1:, :]
#         diff = len(temp_df)-len(df)

#         if(diff != 0):
#             print("WARNING: One or more column(s) had the same nexus hierarchy and only the first was kept.\nThis can happen if your file has merged rows.")
#             print("     Sheet:" + sheet)
#             print("     Number of rows dropped:" + str(diff))

#         df = cleanDF(df)
#         sheet_dict = df.to_dict('index')
#         sheet_dict = {
#             outer_k: {
#                 inner_k: inner_v
#                 for inner_k, inner_v in outer_v.items()
#                 if inner_v != 0 and inner_v != "" and inner_k != "Exists"
#             }
#             for outer_k, outer_v in sheet_dict.items()
#         }

#         for key in sheet_dict:
#             if "Dim" in sheet_dict.get(key):
#                 dim = sheet_dict.get(key).pop("Dim")
#                 rank = sheet_dict.get(key).pop("Rank")
#                 if dim != 0:
#                     dim = "[" + str(dim) + "]"
#                     newitem = {"dimensions": {"dim": dim, "rank": rank}}
#                     sheet_dict.get(key).update(newitem)

#         # pprint.pprint(sheet_dict)
#         output_path = os.path.join(folder_path, yaml_filename)
#         output = {}
#         with open("Headers.yml", 'r') as stream:
#             try:
#                 parsed_yaml = yaml.safe_load(stream)
#                 if(sheet in parsed_yaml):
#                     if 'doc' in parsed_yaml[sheet]:
#                         parsed_yaml[sheet].update({'doc': "\"" + str(parsed_yaml[sheet]['doc']) + "\""})
#                     output.update(parsed_yaml[sheet])
#                     output['(NXobject)'] = sheet_dict
#                 else:
#                     print("ERROR: Not found in Headers.yml; defaulting to no header information.")
#                     tempstring = "(" + sheet + ")"
#                     output[tempstring] = sheet_dict
#             except yaml.YAMLError as exc:
#                 print(exc)
#         with open(output_path, "w") as f:
#             yaml.dump(output, f, sort_keys=False)
#             f.close()

#         clean_duplicate_handler(output_path)

#     xls.close()
#     print("The new file has been saved as " + yaml_filename)


def to_xlsx():
    # print("to be done")
    with open(file_path, 'r') as stream:
        try:
            parsed_yaml = yaml.safe_load(stream)
            # pprint.pprint(parsed_yaml)
            for key in parsed_yaml:
                print(key)
                xlsx_filename = str(key).replace("(", "").replace(")", "") + "_converted.xlsx"
                outputfile = os.path.join(folder_path, xlsx_filename)
                writer = ExcelWriter(outputfile)
                one_sheet = parsed_yaml[key]
                holder = []
                names = []
                types = []
                for row in one_sheet:
                    values = one_sheet[row]
                    df = pd.DataFrame(data=values, index=[0])
                    # print("asdfasdf")
                    nexusheirarchy = row.split(" ")[0]
                    type = row.split(" ", 1)[1].replace("(", "").replace(")", "")
                    # print(nexusheirarchy)
                    # print(type)
                    # print(df)
                    # print(names)
                    holder.append(df)
                    names.append(nexusheirarchy)
                    types.append(type)
                output = pd.concat(holder)
                output.insert(loc=0, column='Nexus heirarchy', value=names)
                output.insert(loc=1, column='Type', value=types)
                output['Dim'] = output['Dim'].fillna(0)
                output['Rank'] = output['Rank'].fillna(0)
                output.to_excel(writer, key, index=False)

            writer.save()
        except yaml.YAMLError as exc:
            print(exc)


def main(args):
    print("Parameter file yaml-tabular parser.")
    if not os.path.exists(args.destination):
        os.makedirs(args.destination)

    args.filename = os.path.join(*args.filename.split("/"))
    # dir_path = os.path.dirname(__file__)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    global folder_path
    global file_path
    folder_path = os.path.join(dir_path, args.destination)
    file_path = os.path.join(dir_path, args.filename)
    print("outputfolder_path: " + folder_path)
    print("readfile_path: " + file_path)

    if args.filename.split(".", 1)[1] == "xlsx":
        xlsx_to_yaml()
    elif args.filename.split(".", 1)[1] == "csv":
        csv_to_yaml()
    elif args.filename.split(".", 1)[1] == "yaml":
        to_xlsx()
    elif args.filename.split(".", 1)[1] == "yml":
        to_xlsx()


if __name__ == "__main__":
    """ This is executed when run from the command line """
    parser = argparse.ArgumentParser()

    # Optional argument which requires a parameter (eg. -d test)
    parser.add_argument("-n", "--filename", action="store", dest="filename", help="filename", required=True)

    parser.add_argument("-d", "--dest", action="store", dest="destination", help="Folder in which to save output files", required=True)

    args = parser.parse_args()
    main(args)
